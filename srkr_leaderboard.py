#chage college name is you need other college details
COLLEGE = 'srkr'
import urllib2 as ul2
from bs4 import BeautifulSoup
import json
from openpyxl import load_workbook
from openpyxl import Workbook

def is_int_then(numb):
    try:
        int(numb)
        return int(numb)
    except Exception as e:
        return numb

r = ul2.urlopen('http://13.126.115.133/')
soup = BeautifulSoup(r.read(),'html.parser')
student_json = soup.find_all('div',attrs = {"data-react-class":"StudentIndexView"});
student_json = student_json[0]['data-react-props']
student_data = json.loads(student_json)
student_data = student_data['studentsData']['data']
wb = Workbook()
ws = wb.active
headers = ['S.No','Name','Gender','lessons_completed','Percentage','Total Score',' C-Basics','C-Arrays-Worksheet','C-Strings-Worksheet','C-LinkedLists-Worksheet','C-Arrays2-Worksheet','C-Strings2-Worksheet','C-LinkedLists2-Worksheet','CodeComplexity','CGames','C-FunctionPointers','C-Recursion','C-BinarySearchTree']

#writing headers
for col in range(0,len(headers)):
    ws.cell(row=1,column=col+1).value = headers[col]

#writing student data    
for student in student_data:
    s_details = []
    if str(student['College']).lower() == COLLEGE:
        try:
            for h in headers:
                if(h=='lessons_completed'):
                    s_details.append('0')
                else:
                    s_details.append(student[h])
            s_details = map(lambda x:is_int_then(x.strip()),s_details)
            lessons = 0
            for marks in s_details[6:]:
                if marks==100:
                    lessons += 1
            s_details[3] = lessons
            ws.append(s_details)
        except Exception as e:
            print str(e)
            pass
wb.save('SRKR_MRND.xlsx')
print 'Done Loading the Srkr college people details......'
